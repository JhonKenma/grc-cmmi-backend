# apps/encuestas/utils.py - VERSIÓN CORREGIDA
import pandas as pd
import openpyxl
from django.db import transaction
from .models import Encuesta, Dimension, Pregunta, NivelReferencia
from rest_framework.exceptions import ValidationError
from .datos_ejemplo_plantilla import DATOS_EJEMPLO_PLANTILLA, NOTA_EXPLICATIVA  # ✅ AGREGAR ESTA LÍNEA


class CargadorExcel:
    """
    Clase para procesar y cargar encuestas desde Excel
    
    Estructura esperada del Excel (VERTICAL - 5 filas por pregunta):
    
    Columnas:
    - seccion_codigo: Código de la dimensión/sección (1, 2, 3)
    - seccion_nombre: Nombre de la dimensión
    - pregunta_codigo: Código de la pregunta (G1.1, D2.1)
    - pregunta_titulo: Título corto de la pregunta
    - pregunta_texto: Texto completo de la pregunta
    - nivel_numero: Número del nivel (1, 2, 3, 4, 5)
    - nivel_descripcion: Descripción de ese nivel
    - nivel_recomendaciones: Recomendaciones para ese nivel
    - nivel_deseado: Nivel objetivo (SOLO en nivel 1)
    - peso: Peso de la pregunta (SOLO en nivel 1)
    
    IMPORTANTE: Cada pregunta ocupa 5 FILAS (una por cada nivel)
    """
    
    COLUMNAS_REQUERIDAS = [
        'seccion_codigo',
        'seccion_nombre',
        'pregunta_codigo',
        'pregunta_titulo',
        'pregunta_texto',
        'nivel_numero',
        'nivel_descripcion',
        'nivel_recomendaciones'
    ]
    
    def __init__(self, archivo_excel, nombre_encuesta, version='1.0', descripcion=''):
        self.archivo = archivo_excel
        self.nombre_encuesta = nombre_encuesta
        self.version = version
        self.descripcion = descripcion
        self.df = None
        self.errores = []
    
    def validar_estructura(self):
        """Valida que el Excel tenga la estructura correcta"""
        try:
            # Leer Excel usando openpyxl (mejor para archivos grandes)
            wb = openpyxl.load_workbook(self.archivo)
            
            # Buscar hoja "ENCUESTA" o la primera hoja disponible
            if "ENCUESTA" in wb.sheetnames:
                ws = wb["ENCUESTA"]
            else:
                ws = wb.active
            
            # Convertir a DataFrame
            data = ws.values
            cols = next(data)
            self.df = pd.DataFrame(data, columns=cols)
            
            # Limpiar nombres de columnas (quitar espacios)
            self.df.columns = self.df.columns.str.strip()
            
            # Validar columnas requeridas
            columnas_faltantes = []
            for col in self.COLUMNAS_REQUERIDAS:
                if col not in self.df.columns:
                    columnas_faltantes.append(col)
            
            if columnas_faltantes:
                raise ValidationError({
                    'archivo': f'Faltan columnas requeridas: {", ".join(columnas_faltantes)}'
                })
            
            # Eliminar filas vacías (donde seccion_codigo es None)
            self.df = self.df[self.df['seccion_codigo'].notna()]
            
            # Validar que no esté vacío
            if self.df.empty:
                raise ValidationError({
                    'archivo': 'El archivo Excel está vacío o no tiene datos válidos'
                })
            
            # Validar tipos de datos
            self._validar_tipos_datos()
            
            # Validar estructura de preguntas (5 niveles)
            self._validar_estructura_preguntas()
            
            return True
            
        except Exception as e:
            if isinstance(e, ValidationError):
                raise e
            raise ValidationError({
                'archivo': f'Error al leer el archivo: {str(e)}'
            })
    
    def _validar_tipos_datos(self):
        """Valida tipos de datos de columnas específicas"""
        errores = []
        
        # Validar que nivel_numero sea 1-5
        try:
            self.df['nivel_numero'] = pd.to_numeric(self.df['nivel_numero'], errors='coerce')
            niveles_invalidos = self.df[
                (self.df['nivel_numero'] < 1) | (self.df['nivel_numero'] > 5)
            ]
            
            if not niveles_invalidos.empty:
                filas = niveles_invalidos.index.tolist()
                errores.append(
                    f'La columna "nivel_numero" debe contener valores 1-5. '
                    f'Filas con problemas: {filas}'
                )
        except Exception as e:
            errores.append(f'Error en columna nivel_numero: {str(e)}')
        
        # Validar campos obligatorios
        campos_obligatorios = {
            'seccion_codigo': 'Código de sección',
            'seccion_nombre': 'Nombre de sección',
            'pregunta_codigo': 'Código de pregunta',
            'pregunta_titulo': 'Título de pregunta',
            'pregunta_texto': 'Texto de pregunta',
            'nivel_descripcion': 'Descripción de nivel'
        }
        
        for campo, nombre_legible in campos_obligatorios.items():
            valores_vacios = self.df[self.df[campo].isna() | (self.df[campo] == '')]
            if not valores_vacios.empty:
                filas = valores_vacios.index.tolist()
                errores.append(
                    f'El campo "{nombre_legible}" tiene valores vacíos en las filas: {filas}'
                )
        
        if errores:
            raise ValidationError({
                'archivo': errores
            })
    
    def _validar_estructura_preguntas(self):
        """
        Valida que cada pregunta tenga exactamente 5 niveles (1, 2, 3, 4, 5)
        """
        errores = []
        
        # Agrupar por pregunta
        preguntas_agrupadas = self.df.groupby('pregunta_codigo')
        
        for codigo_pregunta, grupo in preguntas_agrupadas:
            # Verificar que tenga 5 niveles
            if len(grupo) != 5:
                errores.append(
                    f'Pregunta {codigo_pregunta} debe tener exactamente 5 niveles, '
                    f'tiene {len(grupo)}'
                )
                continue
            
            # Verificar que sean niveles 1, 2, 3, 4, 5
            niveles = sorted(grupo['nivel_numero'].tolist())
            if niveles != [1, 2, 3, 4, 5]:
                errores.append(
                    f'Pregunta {codigo_pregunta} debe tener niveles 1-5, '
                    f'tiene: {niveles}'
                )
        
        if errores:
            raise ValidationError({
                'estructura': errores
            })
    
    @transaction.atomic
    def procesar_y_guardar(self):
        """
        Procesa el DataFrame y guarda en la base de datos
        Retorna la encuesta creada
        """
        # Validar estructura primero
        self.validar_estructura()
        
        print(f"\n{'='*60}")
        print(f"📊 INICIANDO CARGA DE ENCUESTA: {self.nombre_encuesta}")
        print(f"{'='*60}\n")
        
        # Crear encuesta
        encuesta = Encuesta.objects.create(
            nombre=self.nombre_encuesta,
            descripcion=self.descripcion,
            version=self.version,
            es_plantilla=True,
            activo=False
        )
        print(f"✅ Encuesta creada: {encuesta.nombre}")
        
        # Diccionarios para cachear objetos creados
        dimensiones_cache = {}
        preguntas_cache = {}
        
        # Procesar fila por fila
        for index, row in self.df.iterrows():
            try:
                # Extraer datos
                seccion_codigo = str(row['seccion_codigo']).strip()
                seccion_nombre = str(row['seccion_nombre']).strip()
                pregunta_codigo = str(row['pregunta_codigo']).strip()
                pregunta_titulo = str(row['pregunta_titulo']).strip()
                pregunta_texto = str(row['pregunta_texto']).strip()
                nivel_numero = int(row['nivel_numero'])
                nivel_descripcion = str(row['nivel_descripcion']).strip()
                nivel_recomendaciones = str(row['nivel_recomendaciones']).strip() if pd.notna(row['nivel_recomendaciones']) else ''
                
                # Peso y nivel_deseado solo en nivel 1
                peso = 1.0
                nivel_deseado = None
                
                if nivel_numero == 1:
                    if 'peso' in row and pd.notna(row['peso']):
                        peso = float(row['peso'])
                    if 'nivel_deseado' in row and pd.notna(row['nivel_deseado']):
                        nivel_deseado = int(row['nivel_deseado'])
                
                # ==========================================
                # 1. CREAR O RECUPERAR DIMENSIÓN
                # ==========================================
                dimension_key = seccion_codigo
                
                if dimension_key not in dimensiones_cache:
                    # Contar dimensiones existentes para orden
                    orden = len(dimensiones_cache) + 1
                    
                    dimension = Dimension.objects.create(
                        encuesta=encuesta,
                        codigo=seccion_codigo,
                        nombre=seccion_nombre,
                        descripcion='',
                        orden=orden,
                        activo=True
                    )
                    dimensiones_cache[dimension_key] = dimension
                    print(f"   📁 Dimensión creada: {dimension.codigo} - {dimension.nombre}")
                else:
                    dimension = dimensiones_cache[dimension_key]
                
                # ==========================================
                # 2. CREAR O RECUPERAR PREGUNTA
                # ==========================================
                pregunta_key = pregunta_codigo
                
                if pregunta_key not in preguntas_cache:
                    # Contar preguntas de esta dimensión para orden
                    orden_pregunta = Pregunta.objects.filter(dimension=dimension).count() + 1
                    
                    pregunta = Pregunta.objects.create(
                        dimension=dimension,
                        codigo=pregunta_codigo,
                        titulo=pregunta_titulo,
                        texto=pregunta_texto,
                        peso=peso,
                        obligatoria=True,
                        orden=orden_pregunta,
                        activo=True
                    )
                    preguntas_cache[pregunta_key] = pregunta
                    print(f"      ❓ Pregunta creada: {pregunta.codigo} - {pregunta.titulo}")
                else:
                    pregunta = preguntas_cache[pregunta_key]
                
                # ==========================================
                # 3. CREAR NIVEL DE REFERENCIA
                # ==========================================
                nivel = NivelReferencia.objects.create(
                    pregunta=pregunta,
                    numero=nivel_numero,
                    descripcion=nivel_descripcion,
                    recomendaciones=nivel_recomendaciones,
                    activo=True
                )
                print(f"         ✓ Nivel {nivel_numero} creado")
                
            except Exception as e:
                error_msg = f'Error en fila {index + 2}: {str(e)}'
                self.errores.append(error_msg)
                print(f"         ❌ {error_msg}")
        
        # Si hubo errores, revertir transacción
        if self.errores:
            print(f"\n❌ ERRORES ENCONTRADOS:")
            for error in self.errores:
                print(f"   - {error}")
            raise ValidationError({
                'procesamiento': self.errores
            })
        
        # Estadísticas finales
        print(f"\n{'='*60}")
        print(f"✅ CARGA COMPLETADA EXITOSAMENTE")
        print(f"{'='*60}")
        print(f"📊 Dimensiones creadas: {len(dimensiones_cache)}")
        print(f"📊 Preguntas creadas: {len(preguntas_cache)}")
        print(f"📊 Niveles creados: {NivelReferencia.objects.filter(pregunta__dimension__encuesta=encuesta).count()}")
        print(f"{'='*60}\n")
        
        return encuesta
    
    @staticmethod
    def generar_plantilla_excel():
        """
        Genera un archivo Excel con las columnas correctas Y DATOS DE EJEMPLO
        """
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ENCUESTA"
        
        # Estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = [
            'seccion_codigo', 'seccion_nombre', 'pregunta_codigo',
            'pregunta_titulo', 'pregunta_texto', 'nivel_numero',
            'nivel_descripcion', 'nivel_recomendaciones',
            'nivel_deseado', 'peso'
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 50
        ws.column_dimensions['H'].width = 70
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 10
        
        # ✅ AGREGAR DATOS DE EJEMPLO
        for row_idx, fila in enumerate(DATOS_EJEMPLO_PLANTILLA, start=2):
            for col_idx, valor in enumerate(fila, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = valor
                cell.fill = example_fill  # Fondo gris
                cell.alignment = left_alignment if col_idx in [2, 4, 5, 7, 8] else center_alignment
                cell.border = thin_border
        
        # ✅ AGREGAR NOTA EXPLICATIVA
        nota_row = len(DATOS_EJEMPLO_PLANTILLA) + 3
        ws.cell(row=nota_row, column=1).value = "NOTA:"
        ws.cell(row=nota_row, column=1).font = Font(bold=True, color="FF0000")
        
        ws.merge_cells(f'B{nota_row}:J{nota_row}')
        nota_cell = ws.cell(row=nota_row, column=2)
        nota_cell.value = NOTA_EXPLICATIVA
        nota_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        nota_cell.font = Font(italic=True, color="0000FF")
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Guardar en BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output