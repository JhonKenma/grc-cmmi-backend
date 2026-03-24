# apps/reportes/exporters_iq.py
"""
Exportadores PDF y Excel para Reportes de Evaluaciones Inteligentes (IQ).
Reutilizan la misma base (BaseExporter) que los exportadores de encuestas.
"""
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, PageBreak,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from apps.reportes.exporters.base import BaseExporter


# ─────────────────────────────────────────────────────────────────────────────
# BASE IQ  (adapta BaseExporter para usar AsignacionEvaluacionIQ)
# ─────────────────────────────────────────────────────────────────────────────

class BaseIQExporter(BaseExporter):
    """
    BaseExporter adaptado para IQ.
    BaseExporter espera self.evaluacion y self.calculos.
    Aquí self.evaluacion = AsignacionEvaluacionIQ.
    """

    def __init__(self, asignacion, calculos):
        # Llamamos al padre con asignacion como "evaluacion"
        super().__init__(asignacion, calculos)
        self.asignacion = asignacion

    # ── Métodos que BaseExporter llama internamente ────────────────────────────

    def get_estadisticas_generales(self):
        from django.db.models import Avg, Count
        stats = self.calculos.aggregate(
            nivel_deseado_avg=Avg('nivel_deseado'),
            nivel_actual_avg=Avg('nivel_actual'),
            gap_avg=Avg('gap'),
            cumplimiento_avg=Avg('porcentaje_cumplimiento'),
        )
        return {
            'nivel_deseado_avg':  float(stats['nivel_deseado_avg'] or 0),
            'nivel_actual_avg':   float(stats['nivel_actual_avg'] or 0),
            'gap_avg':            float(stats['gap_avg'] or 0),
            'cumplimiento_avg':   float(stats['cumplimiento_avg'] or 0),
            'total_dimensiones':  self.calculos.values('seccion').distinct().count(),
            'total_usuarios':     1,
        }

    def get_dimensiones_data(self):
        """Devuelve shape compatible con ChartGenerator (usa dimension.codigo/nombre)."""
        from collections import namedtuple
        Dim = namedtuple('Dim', ['codigo', 'nombre'])

        resultado = []
        for calculo in self.calculos.order_by('framework_id', 'seccion'):
            codigo = calculo.seccion[:10].upper().replace(' ', '_')
            resultado.append({
                'dimension':         Dim(codigo=codigo, nombre=calculo.seccion),
                'nivel_deseado':     float(calculo.nivel_deseado),
                'nivel_actual_avg':  float(calculo.nivel_actual),
                'gap_avg':           float(calculo.gap),
                'cumplimiento_avg':  float(calculo.porcentaje_cumplimiento),
                'total_usuarios':    1,
            })
        return resultado

    def get_usuarios_data(self):
        usuario = self.asignacion.usuario_asignado
        from django.db.models import Avg, Count
        stats = self.calculos.aggregate(
            nivel_actual_avg=Avg('nivel_actual'),
            gap_avg=Avg('gap'),
            cumplimiento_avg=Avg('porcentaje_cumplimiento'),
        )
        from collections import namedtuple
        User = namedtuple('User', ['nombre_completo', 'email', 'cargo'])
        return [{
            'usuario': User(
                nombre_completo=usuario.get_full_name(),
                email=usuario.email,
                cargo=getattr(usuario, 'cargo', 'N/A'),
            ),
            'nivel_actual_avg':  float(stats['nivel_actual_avg'] or 0),
            'gap_avg':           float(stats['gap_avg'] or 0),
            'cumplimiento_avg':  float(stats['cumplimiento_avg'] or 0),
            'total_dimensiones': self.calculos.count(),
        }]

    def get_clasificaciones_gap(self):
        total = self.calculos.count() or 1
        counts = {
            'critico':  self.calculos.filter(clasificacion_gap='critico').count(),
            'alto':     self.calculos.filter(clasificacion_gap='alto').count(),
            'medio':    self.calculos.filter(clasificacion_gap='medio').count(),
            'bajo':     self.calculos.filter(clasificacion_gap='bajo').count(),
            'cumplido': self.calculos.filter(clasificacion_gap='cumplido').count(),
            'superado': self.calculos.filter(clasificacion_gap='superado').count(),
        }
        for key in list(counts.keys()):
            counts[f'{key}_porcentaje'] = round(counts[key] / total * 100, 1)
        return counts


# ─────────────────────────────────────────────────────────────────────────────
# PDF IQ
# ─────────────────────────────────────────────────────────────────────────────

class PDFExporterIQ(BaseIQExporter):
    """Exporta el reporte de una AsignacionEvaluacionIQ a PDF."""

    def get_content_type(self):
        return 'application/pdf'

    def get_file_extension(self):
        return 'pdf'

    def generate(self):
        doc = SimpleDocTemplate(
            self.buffer,
            pagesize=A4,
            rightMargin=2*cm, leftMargin=2*cm,
            topMargin=1.5*cm, bottomMargin=2*cm,
        )
        self._setup_styles()
        story = []
        story.extend(self._portada())
        story.append(PageBreak())
        story.extend(self._resumen_ejecutivo())
        story.extend(self._seccion_brechas())
        story.append(PageBreak())
        story.extend(self._detalle_secciones())
        story.extend(self._recomendaciones())
        story.extend(self._footer())
        doc.build(story, onFirstPage=self._paginacion, onLaterPages=self._paginacion)

    def _paginacion(self, canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(A4[0] - 2*cm, 1*cm, f"Página {canvas.getPageNumber()}")
        canvas.restoreState()

    def _setup_styles(self):
        self.title_style = ParagraphStyle(
            'T', fontSize=18, textColor=colors.HexColor('#1F4788'),
            fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=8,
        )
        self.h1 = ParagraphStyle(
            'H1', fontSize=13, textColor=colors.HexColor('#1F4788'),
            fontName='Helvetica-Bold', spaceAfter=8, spaceBefore=12,
            backColor=colors.HexColor('#E8EAF6'), leftIndent=8, borderPadding=6,
        )
        self.h2 = ParagraphStyle(
            'H2', fontSize=11, textColor=colors.HexColor('#1F4788'),
            fontName='Helvetica-Bold', spaceAfter=5, spaceBefore=8,
        )
        self.normal = ParagraphStyle(
            'N', fontSize=9, spaceAfter=5, alignment=TA_JUSTIFY, leading=12,
        )
        self.label = ParagraphStyle(
            'L', fontSize=10, textColor=colors.HexColor('#1F4788'),
            fontName='Helvetica-Bold', spaceAfter=3,
        )
        self.value = ParagraphStyle(
            'V', fontSize=10, textColor=colors.HexColor('#424242'),
            fontName='Helvetica', spaceAfter=7,
        )
        self.caption = ParagraphStyle(
            'C', fontSize=8, textColor=colors.grey,
            alignment=TA_CENTER, fontName='Helvetica-Oblique', spaceAfter=6,
        )
        self.footer_style = ParagraphStyle(
            'F', fontSize=10, textColor=colors.HexColor('#1F4788'),
            alignment=TA_CENTER, fontName='Helvetica-Bold',
        )

    def _table_style(self):
        return TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#1F4788')),
            ('TEXTCOLOR',     (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, 0), 9),
            ('ALIGN',         (0, 0), (-1, 0), 'CENTER'),
            ('FONTSIZE',      (0, 1), (-1, -1), 8),
            ('GRID',          (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ])

    def _portada(self):
        a = self.asignacion
        elems = []
        elems.append(Paragraph("🛡️ <b>ShieldGrid</b>", self.label))
        elems.append(Spacer(1, 2*cm))
        elems.append(Paragraph("REPORTE DE EVALUACIÓN INTELIGENTE", self.title_style))
        elems.append(Paragraph(a.evaluacion.nombre, self.caption))
        elems.append(Spacer(1, 1.5*cm))

        elems.append(Paragraph("EMPRESA", self.label))
        elems.append(Paragraph(a.empresa.nombre, self.value))

        elems.append(Paragraph("EVALUADO", self.label))
        elems.append(Paragraph(a.usuario_asignado.get_full_name(), self.value))
        elems.append(Paragraph(a.usuario_asignado.email, self.value))

        elems.append(Paragraph("FRAMEWORKS EVALUADOS", self.label))
        fws = ", ".join(fw.nombre for fw in a.evaluacion.frameworks.all())
        elems.append(Paragraph(fws or 'N/A', self.value))

        elems.append(Paragraph("NIVEL DESEADO", self.label))
        elems.append(Paragraph(
            f"{a.evaluacion.nivel_deseado} — {a.evaluacion.get_nivel_deseado_display()}",
            self.value
        ))

        elems.append(Spacer(1, 0.8*cm))
        elems.append(Paragraph("FECHAS", self.label))
        elems.append(Paragraph(
            f"Fecha límite: {a.fecha_limite.strftime('%d/%m/%Y')} | "
            f"Auditada: {a.fecha_auditada.strftime('%d/%m/%Y') if a.fecha_auditada else 'N/A'}",
            self.value
        ))
        if a.auditado_por:
            elems.append(Paragraph(f"Auditado por: {a.auditado_por.get_full_name()}", self.value))

        elems.append(Spacer(1, 1*cm))
        elems.append(Paragraph(
            f"<b>Generado:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            self.caption
        ))
        return elems

    def _resumen_ejecutivo(self):
        elems = []
        elems.append(Paragraph("1. RESUMEN EJECUTIVO", self.h1))
        stats = self.get_estadisticas_generales()

        data = [
            ['MÉTRICA', 'VALOR', 'INTERPRETACIÓN'],
            ['GAP Promedio',           f"{stats['gap_avg']:.2f}",        self._interpretar_gap(stats['gap_avg'])],
            ['Nivel Actual Promedio',  f"{stats['nivel_actual_avg']:.2f}", f"Sobre {stats['nivel_deseado_avg']:.1f} deseado"],
            ['% Cumplimiento',         f"{stats['cumplimiento_avg']:.1f}%", self._interpretar_cumplimiento(stats['cumplimiento_avg'])],
            ['Secciones Evaluadas',    str(stats['total_dimensiones']),   'Secciones/dominios analizados'],
        ]
        t = Table(data, colWidths=[4*cm, 3.5*cm, 9.5*cm])
        t.setStyle(self._table_style())
        elems.append(t)
        elems.append(Spacer(1, 0.5*cm))
        return elems

    def _seccion_brechas(self):
        """Tabla de brechas ordenadas por prioridad."""
        elems = []
        elems.append(Paragraph("2. BRECHAS IDENTIFICADAS", self.h1))
        elems.append(Paragraph(
            "Las siguientes secciones presentan brechas que requieren atención. "
            "Están ordenadas por prioridad de remediación.",
            self.normal
        ))
        elems.append(Spacer(1, 0.3*cm))

        brechas_qs = self.calculos.filter(
            clasificacion_gap__in=['critico', 'alto', 'medio', 'bajo']
        ).order_by('-gap')

        if not brechas_qs.exists():
            elems.append(Paragraph("✅ No se identificaron brechas significativas.", self.normal))
            return elems

        data = [['PRIORIDAD', 'SECCIÓN', 'FRAMEWORK', 'GAP', 'NIVEL ACTUAL', 'CLASIFICACIÓN']]
        prioridad_map = {'critico': '🔴 CRÍTICO', 'alto': '🟠 ALTO', 'medio': '🟡 MEDIO', 'bajo': '🔵 BAJO'}

        for calculo in brechas_qs:
            data.append([
                prioridad_map.get(calculo.clasificacion_gap, calculo.clasificacion_gap),
                calculo.seccion[:30],
                calculo.framework_nombre[:20],
                f"{calculo.gap:.1f}",
                f"{calculo.nivel_actual:.1f} / {calculo.nivel_deseado:.1f}",
                calculo.get_clasificacion_gap_display(),
            ])

        t = Table(data, colWidths=[2.5*cm, 5*cm, 3.5*cm, 1.5*cm, 2.5*cm, 2*cm])
        t.setStyle(self._table_style())
        elems.append(t)
        elems.append(Spacer(1, 0.5*cm))
        return elems

    def _detalle_secciones(self):
        """Tabla completa de todas las secciones."""
        elems = []
        elems.append(Paragraph("3. DETALLE POR SECCIÓN", self.h1))

        data = [['SECCIÓN', 'FRAMEWORK', 'DESEADO', 'ACTUAL', 'GAP', '% CUMPL.', 'ESTADO']]
        for calculo in self.calculos.order_by('framework_id', 'seccion'):
            estado = '✅ Cumplido' if calculo.gap <= 0 else f"⚠️ {calculo.get_clasificacion_gap_display()}"
            data.append([
                calculo.seccion[:28],
                calculo.framework_nombre[:18],
                f"{calculo.nivel_deseado:.1f}",
                f"{calculo.nivel_actual:.1f}",
                f"{calculo.gap:.1f}",
                f"{calculo.porcentaje_cumplimiento:.0f}%",
                estado,
            ])

        t = Table(data, colWidths=[4.5*cm, 3*cm, 1.8*cm, 1.8*cm, 1.5*cm, 2*cm, 2.4*cm])
        t.setStyle(self._table_style())
        elems.append(t)
        elems.append(Spacer(1, 0.5*cm))
        return elems

    def _recomendaciones(self):
        elems = []
        elems.append(Paragraph("4. RECOMENDACIONES", self.h1))
        stats = self.get_estadisticas_generales()

        recomendaciones = []
        if stats['gap_avg'] >= 2:
            recomendaciones.append(
                "Desarrollar un plan de remediación urgente para las secciones críticas, "
                "asignando responsables y fechas límite concretas."
            )
        if stats['cumplimiento_avg'] < 70:
            recomendaciones.append(
                "Implementar capacitaciones específicas en los frameworks evaluados "
                "para mejorar el nivel de cumplimiento general."
            )

        brechas_criticas = self.calculos.filter(clasificacion_gap__in=['critico', 'alto']).count()
        if brechas_criticas > 0:
            elems_criticos = ", ".join(
                c.seccion[:25] for c in
                self.calculos.filter(clasificacion_gap__in=['critico', 'alto'])[:3]
            )
            recomendaciones.append(
                f"Priorizar las secciones: {elems_criticos}. "
                "Iniciar proyectos de remediación en los próximos 30 días."
            )

        recomendaciones.append(
            "Documentar los controles existentes y establecer métricas de seguimiento "
            "trimestral para monitorear el avance en el cierre de brechas."
        )

        for idx, rec in enumerate(recomendaciones, 1):
            elems.append(Paragraph(f"<b>{idx}.</b> {rec}", self.normal))
            elems.append(Spacer(1, 0.25*cm))

        return elems

    def _footer(self):
        elems = []
        elems.append(Spacer(1, 1.5*cm))
        line = Table([['']], colWidths=[17*cm])
        line.setStyle(TableStyle([('LINEABOVE', (0, 0), (-1, 0), 2, colors.HexColor('#1F4788'))]))
        elems.append(line)
        elems.append(Spacer(1, 0.3*cm))
        elems.append(Paragraph("🛡️ <b>Powered by ShieldGrid</b>", self.footer_style))
        return elems

    def _interpretar_gap(self, gap):
        if gap >= 3: return "⚠️ Crítico — Acción inmediata"
        if gap >= 2: return "⚠️ Alto — Atención prioritaria"
        if gap >= 1: return "⚡ Medio — Plan de mejora"
        if gap > 0:  return "✓ Bajo — Mejoras incrementales"
        return "✓✓ Sin brecha"

    def _interpretar_cumplimiento(self, pct):
        if pct >= 90: return "✓✓ Excelente"
        if pct >= 75: return "✓ Bueno"
        if pct >= 60: return "⚡ Aceptable"
        return "⚠️ Requiere atención"


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL IQ
# ─────────────────────────────────────────────────────────────────────────────

class ExcelExporterIQ(BaseIQExporter):
    """Exporta el reporte de una AsignacionEvaluacionIQ a Excel."""

    def get_content_type(self):
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def get_file_extension(self):
        return 'xlsx'

    def generate(self):
        wb = Workbook()
        self.hfill  = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        self.hfont  = Font(color="FFFFFF", bold=True, size=11)
        self.tfont  = Font(bold=True, size=13)
        self.center = Alignment(horizontal="center", vertical="center")

        self._hoja_resumen(wb)
        self._hoja_secciones(wb)
        self._hoja_brechas(wb)
        self._hoja_respuestas(wb)

        wb.save(self.buffer)

    def _header_row(self, ws, headers):
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.hfill
            cell.font = self.hfont
            cell.alignment = self.center

    def _hoja_resumen(self, wb):
        ws = wb.active
        ws.title = "Resumen"
        a = self.asignacion

        ws['A1'] = "REPORTE EVALUACIÓN INTELIGENTE (IQ)"
        ws['A1'].font = self.tfont
        ws.merge_cells('A1:C1')

        ws['A2'] = f"Evaluación: {a.evaluacion.nombre}"
        ws['A3'] = f"Empresa: {a.empresa.nombre}"
        ws['A4'] = f"Evaluado: {a.usuario_asignado.get_full_name()}"
        ws['A5'] = f"Auditada: {a.fecha_auditada.strftime('%d/%m/%Y') if a.fecha_auditada else 'N/A'}"

        ws['A7'] = "MÉTRICA"
        ws['B7'] = "VALOR"
        ws['A7'].fill = self.hfill; ws['A7'].font = self.hfont
        ws['B7'].fill = self.hfill; ws['B7'].font = self.hfont

        stats = self.get_estadisticas_generales()
        metricas = [
            ("Nivel Deseado Promedio",     stats['nivel_deseado_avg']),
            ("Nivel Actual Promedio",      stats['nivel_actual_avg']),
            ("GAP Promedio",               stats['gap_avg']),
            ("% Cumplimiento Promedio",    stats['cumplimiento_avg']),
            ("Total Secciones Evaluadas",  stats['total_dimensiones']),
        ]
        for i, (k, v) in enumerate(metricas, start=8):
            ws[f'A{i}'] = k
            ws[f'B{i}'] = round(v, 2) if isinstance(v, float) else v

        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 18

    def _hoja_secciones(self, wb):
        ws = wb.create_sheet("Secciones")
        self._header_row(ws, [
            "Sección", "Framework", "Nivel Deseado", "Nivel Actual",
            "GAP", "% Cumplimiento", "Clasificación",
            "Sí Cumple", "Cumple Parcial", "No Cumple", "No Aplica",
        ])
        for c in self.calculos.order_by('framework_id', 'seccion'):
            ws.append([
                c.seccion, c.framework_nombre,
                float(c.nivel_deseado), float(c.nivel_actual),
                float(c.gap), float(c.porcentaje_cumplimiento),
                c.get_clasificacion_gap_display(),
                c.respuestas_si_cumple, c.respuestas_cumple_parcial,
                c.respuestas_no_cumple, c.respuestas_no_aplica,
            ])
        for col, w in zip('ABCDEFGHIJK', [35, 20, 14, 14, 8, 16, 18, 10, 14, 10, 10]):
            ws.column_dimensions[col].width = w

    def _hoja_brechas(self, wb):
        ws = wb.create_sheet("Brechas Identificadas")
        self._header_row(ws, [
            "Prioridad", "Sección", "Framework",
            "Nivel Deseado", "Nivel Actual", "GAP",
            "Clasificación", "% Cumplimiento",
            "Preguntas No Cumple", "Total Preguntas",
        ])
        prioridad_map = {'critico': 1, 'alto': 2, 'medio': 3, 'bajo': 4}
        brechas = self.calculos.filter(
            clasificacion_gap__in=['critico', 'alto', 'medio', 'bajo']
        ).order_by('-gap')

        for c in brechas:
            ws.append([
                prioridad_map.get(c.clasificacion_gap, 5),
                c.seccion, c.framework_nombre,
                float(c.nivel_deseado), float(c.nivel_actual), float(c.gap),
                c.get_clasificacion_gap_display(),
                float(c.porcentaje_cumplimiento),
                c.respuestas_no_cumple, c.total_preguntas,
            ])
        for col, w in zip('ABCDEFGHIJ', [10, 35, 20, 14, 14, 8, 18, 16, 20, 14]):
            ws.column_dimensions[col].width = w

    def _hoja_respuestas(self, wb):
        ws = wb.create_sheet("Distribución Respuestas")
        from django.db.models import Sum
        totales = self.calculos.aggregate(
            si=Sum('respuestas_si_cumple'),
            parcial=Sum('respuestas_cumple_parcial'),
            no=Sum('respuestas_no_cumple'),
            na=Sum('respuestas_no_aplica'),
        )
        self._header_row(ws, ["Tipo de Respuesta", "Cantidad", "Porcentaje"])

        total = sum(v or 0 for v in totales.values()) or 1
        for label, key in [
            ("Sí Cumple", 'si'), ("Cumple Parcial", 'parcial'),
            ("No Cumple", 'no'), ("No Aplica", 'na'),
        ]:
            val = totales[key] or 0
            ws.append([label, val, f"{round(val/total*100, 1)}%"])

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 14